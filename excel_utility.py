import datetime
import pandas as pd
import openpyxl

from openpyxl import load_workbook

''''
This script is for performing CRUD operation in excel
'''

'''
This function is used to read the excel
@return list of dictionary
'''
def read_excel(excel_name='pending_jobs.xlsx',sheet_name='Sheet1'):
    try:
        df = pd.read_excel(excel_name, sheet_name,engine = 'openpyxl')
        # df=df.fillna("")

        data=df.to_dict('records')
        print("The length of data in excel is : ", len(data))
            
            
        return data
    except Exception as e:
        print("Exception encountered in reading excel", e)

'''
This function is used to insert the data in excel
'''
def insert_data_in_excel(df,excel_name='pending_jobs.xlsx', sheet_name='Sheet1', header=True, start_row=0):

    # Create a Pandas Excel writer using XlsxWriter as the engine.
    writer = pd.ExcelWriter(excel_name, engine='openpyxl')

    # Convert the dataframe to an XlsxWriter Excel object.
    df.to_excel(writer, sheet_name=sheet_name, index=False, header=header,startrow=start_row)

    # Close the Pandas Excel writer and output the Excel file.
    writer.close()

'''
This function is used to add the data in excel
'''
def add_to_excel(df, excel_name='pending_jobs.xlsx'):
    # Load workbook
    wb = openpyxl.load_workbook(excel_name)
    
    # Get the active worksheet
    ws = wb.active
    
    # Get the last row in the worksheet
    last_row = ws.max_row
    
    columns={}
    i=1
    for cell in ws[1]:
        columns[cell.value] = i
        i=i+1 
    print(df.columns) 

    # Iterate through the DataFrame and add the rows to the worksheet
    for index, row in df.iterrows():
        for col, value in row.items():
            ws.cell(row=last_row+index+1, column=columns[col], value=value)
            i=i+1
    # Save the workbook
    wb.save(excel_name)

'''
This function is used to delete the data from excel
'''
def delete_from_excel(filepath='pending_jobs.xlsx', column_name=None, filter_value=None):
    # Load workbook
    if not column_name: 
        return
    wb = openpyxl.load_workbook(filepath)
    
    # Get the active worksheet
    ws = wb.active
    
    # Get the names of the columns
    columns = []
    for cell in ws[1]:
        columns.append(cell.value)

    print("Columns:" , columns)
    
    # Get the index of the column with the specified name
    column_index = columns.index(column_name) + 1
    print("Column_index:" , column_index)
    
    # Get the rows to delete
    rows_to_delete = []
    for row in ws.iter_rows():
        if row[column_index-1].value == filter_value:
            rows_to_delete.append(row[0].row)
    
    # Delete the rows
    print("Rows_to_delete: ", rows_to_delete)
    for row in rows_to_delete[::-1]:
        # print([ print(each.value) for each in ws[row] ])
        ws.delete_rows(row, 1)
    
    # Save the workbook
    wb.save(filepath)

if __name__=='__main__':

    each_rec={'ORDER_ID': '2ws5i', 'ORDER_DATE': datetime.datetime(2023, 1, 6, 0, 0), 'JOB_ID': '0', 'JOB_MEM_NAME': 'NRT_API_TIEOUT_GPSS_PROD', 'START_TIME': datetime.datetime(2023, 1, 6, 3, 1, 1), 'END_TIME': datetime.datetime(2023, 1, 6, 3, 1, 1), 'RERUN_COUNTER': 13.0, 'ENDED_STATUS': 32.0}
    l=[each_rec,each_rec,each_rec, each_rec]

    # df = pd.DataFrame(each_rec, index=[0])

    df = pd.DataFrame.from_dict(l)

    insert_data_in_excel(df)

    each_rec={'ORDER_ID': '2wskk', 'ORDER_DATE': datetime.datetime(2023, 1, 6, 0, 0), 'JOB_ID': '0', 'JOB_MEM_NAME': 'NRT_API_TIEOUT_GPSS_PROD', 'START_TIME': datetime.datetime(2023, 1, 6, 3, 1, 1), 'END_TIME': datetime.datetime(2023, 1, 6, 3, 1, 1), 'RERUN_COUNTER': 13.0, 'ENDED_STATUS': 32.0}
    l=[each_rec,each_rec,each_rec, each_rec]
    df = pd.DataFrame.from_dict(l)

    add_to_excel(df)

    each_rec={'ORDER_ID': '2wskfff', 'ORDER_DATE': datetime.datetime(2023, 1, 6, 0, 0), 'JOB_ID': '0', 'JOB_MEM_NAME': 'NRT_API_TIEOUT_GPSS_PROD', 'START_TIME': datetime.datetime(2023, 1, 6, 3, 1, 1), 'END_TIME': datetime.datetime(2023, 1, 6, 3, 1, 1), 'RERUN_COUNTER': 13.0, 'ENDED_STATUS': 32.0}
    l=[each_rec, each_rec]
    df = pd.DataFrame.from_dict(l)

    add_to_excel(df)

    delete_from_excel(column_name='ORDER_ID', filter_value='2wskfff')

    read_excel()


